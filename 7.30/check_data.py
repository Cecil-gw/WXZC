# -*- coding: utf-8 -*-
"""快速检查 data.xlsx 文件格式"""
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl 未安装")
    sys.exit(1)

wb = openpyxl.load_workbook(r"d:\wx26.7.14\7.30\data\data.xlsx", read_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row - 1}  (不含表头)")
print(f"Cols: {ws.max_column}")

headers = [cell.value for cell in ws[1]]
print(f"\nHeaders ({len(headers)} 列):")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# 检查必需列
required = [
    "id", "Gender", "Age", "Driving_License", "Region_Code",
    "Previously_Insured", "Vehicle_Age", "Vehicle_Damage",
    "Annual_Premium", "Policy_Sales_Channel", "Vintage", "Response"
]

actual = set(headers)
missing = [c for c in required if c not in actual]
extra = [c for c in headers if c not in required]

if missing:
    print(f"\n⚠️  缺少必需列: {missing}")
else:
    print(f"\n✅ 所有必需列齐全")

if extra:
    print(f"ℹ️  额外列: {extra}")

# 看前 3 行数据
print("\n前 3 行数据:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f"  Row {i+1}: {row}")

wb.close()
